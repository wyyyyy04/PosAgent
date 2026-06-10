"""
Excel 写入模块 — 将匹配结果写入模板 Excel，保留原始格式，生成校验报告。
"""

from pathlib import Path
from typing import List, Optional

import pandas as pd
import openpyxl

# 在输出文件中新增的置信度列名
CONFIDENCE_COLUMN = "匹配置信度"

# 置信度值
HIGH = "HIGH"
LOW_CONFIDENCE = "LOW_CONFIDENCE"


def write_result(
    template_path: str,
    output_path: str,
    result_df: pd.DataFrame,
    target_col: str = "配料",
    confidence_col: str = CONFIDENCE_COLUMN,
    header_row: int = 1,
) -> str:
    """将匹配结果写入 Excel，保留模板原始格式。

    以 openpyxl 打开模板文件，只修改目标列和置信度列，
    其他单元格的样式、公式、数据验证均保持不变。

    Args:
        template_path: 原始模板 Excel 路径。
        output_path: 输出文件路径。
        result_df: 包含填充结果的 DataFrame，须至少包含 target_col 和 confidence_col。
        target_col: 需要填充的目标列名，默认 "配料"。
        confidence_col: 置信度列名，默认 "匹配置信度"。
        header_row: 表头所在行号（1=第一行），chowbus 模板为 1（两行表头，数据从第3行开始）。

    Returns:
        output_path
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 定位目标列和置信度列的列号
    target_col_idx = None
    confidence_col_idx = None
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip() == target_col:
            target_col_idx = col
        if val and str(val).strip() == confidence_col:
            confidence_col_idx = col

    if target_col_idx is None:
        # 目标列不存在，追加到末尾
        target_col_idx = max_col + 1
        if header_row == 1:
            ws.cell(row=1, column=target_col_idx, value=target_col)
        else:
            # 多行表头：仅在第一行写入列名
            ws.cell(row=1, column=target_col_idx, value=target_col)

    if confidence_col_idx is None:
        # 置信度列不存在，追加到目标列后面
        confidence_col_idx = target_col_idx + 1
        if header_row == 1:
            ws.cell(row=1, column=confidence_col_idx, value=confidence_col)
        else:
            ws.cell(row=1, column=confidence_col_idx, value=confidence_col)

    # 写入数据（从表头下一行开始，chowbus 双表头从第3行开始）
    data_start_row = header_row + 1
    for i, (_, row) in enumerate(result_df.iterrows()):
        excel_row = data_start_row + i
        if target_col in result_df.columns:
            ws.cell(row=excel_row, column=target_col_idx, value=row.get(target_col, ""))
        if CONFIDENCE_COLUMN in result_df.columns:
            ws.cell(row=excel_row, column=confidence_col_idx, value=row.get(CONFIDENCE_COLUMN, ""))

    wb.save(output_path)
    return output_path


def write_report(report_path: str, low_confidence_rows: List[dict]) -> str:
    """生成校验报告，汇总所有低置信度行及失败原因。

    Args:
        report_path: 报告输出路径。
        low_confidence_rows: 低置信度行列表，每项为 dict，包含 row_index、reason 等。

    Returns:
        report_path
    """
    lines = [
        "=" * 60,
        "POS Template Mapping — 校验报告",
        "=" * 60,
        "",
    ]

    if not low_confidence_rows:
        lines.append("全部行匹配置信度为 HIGH，无需关注。")
    else:
        lines.append(f"低置信度行数: {len(low_confidence_rows)}")
        lines.append("")
        lines.append("-" * 60)
        for item in low_confidence_rows:
            row_idx = item.get("row_index", "?")
            reason = item.get("reason", "未知原因")
            product = item.get("product_name", "?")
            lines.append(f"  行 {row_idx}: {product}")
            lines.append(f"    原因: {reason}")
            lines.append("")
        lines.append("-" * 60)

    lines.append("")
    lines.append(f"报告生成完毕。")

    Path(report_path).write_text("\n".join(lines), encoding="utf-8")
    return report_path


# ── 自测 ──────────────────────────────────────────────────────

if __name__ == "__main__":
    import tempfile
    import os

    passed = 0
    failed = 0

    def check(condition, msg):
        global passed, failed
        if condition:
            passed += 1
            print(f"  PASS  {msg}")
        else:
            failed += 1
            print(f"  FAIL  {msg}")

    print("=== Excel Writer 自测 ===\n")

    tmpdir = tempfile.mkdtemp()

    # ── 准备模板文件和结果数据 ──
    template_path = os.path.join(tmpdir, "template.xlsx")
    output_path = os.path.join(tmpdir, "output.xlsx")
    report_path = os.path.join(tmpdir, "report.txt")

    # 使用 openpyxl 直接创建带样式的模板（验证样式保留）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = ["菜品名称", "规格", "口味做法组合", "配料"]
    data = [
        ["五黄高纤慢养瓶", "五角瓶", "红茶, 十二分糖, 温热", ""],
        ["五黄高纤慢养瓶", "五角瓶", "红茶, 十二分糖, 正常冰", ""],
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    ws.column_dimensions["A"].width = 20
    wb.save(template_path)

    # 构造结果 DataFrame
    result_df = pd.DataFrame({
        "配料":       ["T240、B30/80、S4", "T265、B30/105、S5"],
        "匹配置信度":  [HIGH, LOW_CONFIDENCE],
    })

    try:
        # ── 1. 写入结果并保留格式 ──
        print("1. 写入结果并保留格式")
        out = write_result(template_path, output_path, result_df)

        # 验证输出文件存在
        check(os.path.exists(out), "输出文件已生成")

        # 读取验证内容
        wb_out = openpyxl.load_workbook(output_path)
        ws_out = wb_out.active

        check(ws_out.cell(row=1, column=1).value == "菜品名称", "表头保留 '菜品名称'")
        check(ws_out.cell(row=1, column=4).value == "配料", "表头保留 '配料'")

        # 验证填充数据
        check(ws_out.cell(row=2, column=4).value == "T240、B30/80、S4", "第1行配料已填充")
        check(ws_out.cell(row=3, column=4).value == "T265、B30/105、S5", "第2行配料已填充")

        # 验证置信度列
        check(ws_out.cell(row=1, column=5).value == "匹配置信度", "置信度列表头已添加")
        check(ws_out.cell(row=2, column=5).value == "HIGH", "第1行 HIGH")
        check(ws_out.cell(row=3, column=5).value == "LOW_CONFIDENCE", "第2行 LOW_CONFIDENCE")

        # 验证样式保留（表头加粗 + 蓝色背景）
        cell_a1 = ws_out.cell(row=1, column=1)
        check(cell_a1.font.bold is True, "表头加粗保留")
        check(cell_a1.fill.fgColor.rgb in ("004472C4", "FF4472C4"), "表头蓝色背景保留")

        # 验证列宽保留
        check(ws_out.column_dimensions["A"].width == 20, "列宽保留")
        print()

        # ── 2. 校验报告 ──
        print("2. 校验报告")
        low_rows = [
            {"row_index": 2, "product_name": "五黄高纤慢养瓶", "reason": "商品名匹配低于阈值"},
        ]
        rp = write_report(report_path, low_rows)
        check(os.path.exists(rp), "报告文件已生成")

        report_text = Path(rp).read_text(encoding="utf-8")
        check("低置信度行数: 1" in report_text, "报告含低置信度计数")
        check("商品名匹配低于阈值" in report_text, "报告含失败原因")
        print()

        # ── 3. 空报告（全部 HIGH） ──
        print("3. 空报告（全部 HIGH）")
        empty_report = os.path.join(tmpdir, "empty_report.txt")
        write_report(empty_report, [])
        empty_text = Path(empty_report).read_text(encoding="utf-8")
        check("全部行匹配置信度为 HIGH" in empty_text, "空报告提示无需关注")
        print()

        # ── 4. 目标列已存在时的覆盖 ──
        print("4. 目标列覆盖写入")
        out2 = write_result(template_path, output_path, result_df)
        wb2 = openpyxl.load_workbook(out2)
        ws2 = wb2.active
        check(ws2.cell(row=2, column=4).value == "T240、B30/80、S4", "覆盖后数据正确")
        print()

    finally:
        for f in [template_path, output_path, report_path,
                  os.path.join(tmpdir, "empty_report.txt")]:
            if os.path.exists(f):
                os.remove(f)
        os.rmdir(tmpdir)

    print(f"=== 结果: {passed} passed, {failed} failed ===")
