"""
Excel 写入模块 — 将匹配结果写入模板 Excel，保留原始格式，生成校验报告。
"""

from pathlib import Path
from typing import List, Optional

import pandas as pd
import openpyxl

# 在输出文件中新增的置信度列名
CONFIDENCE_COLUMN = "匹配置信度"

# 置信度值
HIGH = "HIGH"
LOW_CONFIDENCE = "LOW_CONFIDENCE"


def write_result(
    template_path: str,
    output_path: str,
    result_df: pd.DataFrame,
    target_col: str = "配料",
    confidence_col: str = CONFIDENCE_COLUMN,
    header_row: int = 1,
    data_start_row: Optional[int] = None,
    sheet_name: int = 0,
) -> str:
    """将匹配结果写入 Excel，保留模板原始格式。

    以 openpyxl 打开模板文件，只修改目标列和置信度列，
    其他单元格的样式、公式、数据验证均保持不变。

    Args:
        template_path: 原始模板 Excel 路径。
        output_path: 输出文件路径。
        result_df: 包含填充结果的 DataFrame，须至少包含 target_col 和 confidence_col。
        target_col: 需要填充的目标列名，默认 "配料"。
        confidence_col: 置信度列名，默认 "匹配置信度"。
        header_row: 用于搜索目标列/置信度列的表头行号（1=第一行）。
        data_start_row: 数据写入起始行号，默认 = header_row + 1。
            chowbus 模板应传 header_row=1, data_start_row=3（跳过两行表头）。
        sheet_name: 目标 Sheet 名称或索引（0=第一个 Sheet）。

    Returns:
        output_path
    """
    if data_start_row is None:
        data_start_row = header_row + 1

    wb = openpyxl.load_workbook(template_path)
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    # 定位目标列和置信度列的列号（始终在 header_row 搜索）
    target_col_idx = None
    confidence_col_idx = None
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip() == target_col:
            target_col_idx = col
        if val and str(val).strip() == confidence_col:
            confidence_col_idx = col

    if target_col_idx is None:
        # 目标列不存在，追加到末尾
        target_col_idx = max_col + 1
        ws.cell(row=header_row, column=target_col_idx, value=target_col)

    if confidence_col_idx is None:
        # 置信度列不存在，追加到目标列后面
        confidence_col_idx = target_col_idx + 1
        ws.cell(row=header_row, column=confidence_col_idx, value=confidence_col)

    # 写入数据
    for i, (_, row) in enumerate(result_df.iterrows()):
        excel_row = data_start_row + i
        if target_col in result_df.columns:
            ws.cell(row=excel_row, column=target_col_idx, value=row.get(target_col, ""))
        if CONFIDENCE_COLUMN in result_df.columns:
            ws.cell(row=excel_row, column=confidence_col_idx, value=row.get(CONFIDENCE_COLUMN, ""))

    wb.save(output_path)
    return output_path


def write_report(report_path: str, low_confidence_rows: List[dict]) -> str:
    """生成校验报告，汇总所有低置信度行及失败原因。

    Args:
        report_path: 报告输出路径。
        low_confidence_rows: 低置信度行列表，每项为 dict，包含 row_index、reason 等。

    Returns:
        report_path
    """
    lines = [
        "=" * 60,
        "POS Template Mapping — 校验报告",
        "=" * 60,
        "",
    ]

    if not low_confidence_rows:
        lines.append("全部行匹配置信度为 HIGH，无需关注。")
    else:
        lines.append(f"低置信度行数: {len(low_confidence_rows)}")
        lines.append("")
        lines.append("-" * 60)
        for item in low_confidence_rows:
            row_idx = item.get("row_index", "?")
            reason = item.get("reason", "未知原因")
            product = item.get("product_name", "?")
            lines.append(f"  行 {row_idx}: {product}")
            lines.append(f"    原因: {reason}")
            lines.append("")
        lines.append("-" * 60)

    lines.append("")
    lines.append(f"报告生成完毕。")

    Path(report_path).write_text("\n".join(lines), encoding="utf-8")
    return report_path


# ── 选项规格模板写入 ──────────────────────────────────────────────

# 选项规格模板的 8 个固定列名
OPTION_TEMPLATE_COLUMNS = [
    "商品编码", "商品名称", "口味做法组名", "选项名称",
    "最少必选", "最多可选", "推荐项", "默认项",
]


def write_expanded_template(
    template_path: str,
    output_path: str,
    expanded_df: pd.DataFrame,
    header_row: int = 1,
) -> str:
    """将展开后的选项规格数据写入模板 Excel，保留原始格式。

    以 openpyxl 打开模板，定位 8 个固定列头，清除表头以下旧数据后写入。

    Args:
        template_path: 空白选项模板 Excel 路径（含表头，无数据行）。
        output_path: 输出文件路径。
        expanded_df: expand_master_to_options() 返回的 DataFrame。
        header_row: 列头所在行号（1=第一行）。

    Returns:
        output_path

    Raises:
        FileNotFoundError: 模板文件不存在。
    """
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(
            f"模板文件不存在: {template_path}\n"
            f"请先用 Excel 创建一个包含以下表头的空白模板文件:\n"
            f"  {', '.join(OPTION_TEMPLATE_COLUMNS)}"
        )

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ── Step 1: 定位 8 个固定列头 ──
    max_col = ws.max_column
    col_map = {}  # column_name → 1-based column index

    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            cleaned = str(val).strip()
            # 去除 * 后缀（真实模板常见：商品编码* → 商品编码）
            if cleaned.endswith("*"):
                cleaned_no_star = cleaned[:-1]
            else:
                cleaned_no_star = cleaned
            if cleaned in OPTION_TEMPLATE_COLUMNS:
                col_map[cleaned] = col
            elif cleaned_no_star in OPTION_TEMPLATE_COLUMNS:
                col_map[cleaned_no_star] = col

    # 缺失的列头追加到末尾
    next_col = max_col + 1
    for col_name in OPTION_TEMPLATE_COLUMNS:
        if col_name not in col_map:
            ws.cell(row=header_row, column=next_col, value=col_name)
            col_map[col_name] = next_col
            next_col += 1

    data_start_row = header_row + 1

    # ── Step 2: 清除表头以下的旧数据 ──
    for row in range(data_start_row, ws.max_row + 1):
        for col_name, col_idx in col_map.items():
            ws.cell(row=row, column=col_idx, value=None)

    # ── Step 3: 写入展开数据 ──
    for i, (_, drow) in enumerate(expanded_df.iterrows()):
        excel_row = data_start_row + i
        for col_name, col_idx in col_map.items():
            if col_name in expanded_df.columns:
                val = drow.get(col_name, "")
                # 最少必选/最多可选 保持整数
                if col_name in ("最少必选", "最多可选"):
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        val = 1
                ws.cell(row=excel_row, column=col_idx, value=val)

    wb.save(output_path)
    return output_path


# ── 自测 ──────────────────────────────────────────────────────

if __name__ == "__main__":
    import tempfile
    import os

    passed = 0
    failed = 0

    def check(condition, msg):
        global passed, failed
        if condition:
            passed += 1
            print(f"  PASS  {msg}")
        else:
            failed += 1
            print(f"  FAIL  {msg}")

    print("=== Excel Writer 自测 ===\n")

    tmpdir = tempfile.mkdtemp()

    # ── 准备模板文件和结果数据 ──
    template_path = os.path.join(tmpdir, "template.xlsx")
    output_path = os.path.join(tmpdir, "output.xlsx")
    report_path = os.path.join(tmpdir, "report.txt")

    # 使用 openpyxl 直接创建带样式的模板（验证样式保留）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = ["菜品名称", "规格", "口味做法组合", "配料"]
    data = [
        ["五黄高纤慢养瓶", "五角瓶", "红茶, 十二分糖, 温热", ""],
        ["五黄高纤慢养瓶", "五角瓶", "红茶, 十二分糖, 正常冰", ""],
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    ws.column_dimensions["A"].width = 20
    wb.save(template_path)

    # 构造结果 DataFrame
    result_df = pd.DataFrame({
        "配料":       ["T240、B30/80、S4", "T265、B30/105、S5"],
        "匹配置信度":  [HIGH, LOW_CONFIDENCE],
    })

    # 创建选项规格模板
    opt_template_path = os.path.join(tmpdir, "opt_template.xlsx")
    opt_output_path = os.path.join(tmpdir, "opt_output.xlsx")
    wb_opt = openpyxl.Workbook()
    ws_opt = wb_opt.active
    for c, h in enumerate(OPTION_TEMPLATE_COLUMNS, 1):
        cell = ws_opt.cell(row=1, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    # 预留一列不相关的列，验证格式保留
    ws_opt.cell(row=1, column=9, value="其他列")
    ws_opt.column_dimensions["A"].width = 15
    wb_opt.save(opt_template_path)

    # 构造展开结果 DataFrame
    expanded_df = pd.DataFrame({
        "商品编码": ["A001"] * 6,
        "商品名称": ["茉莉绿茶"] * 6,
        "口味做法组名": ["糖度", "糖度", "糖度", "温度", "温度", "温度"],
        "选项名称": ["七分糖", "五分糖", "三分糖", "正常冰", "少冰", "去冰"],
        "最少必选": [1] * 6,
        "最多可选": [1] * 6,
        "推荐项": ["是", "否", "否", "是", "否", "否"],
        "默认项": ["否", "是", "否", "否", "是", "否"],
    })

    try:
        # ── 1. 写入结果并保留格式 ──
        print("1. 写入结果并保留格式")
        out = write_result(template_path, output_path, result_df)

        # 验证输出文件存在
        check(os.path.exists(out), "输出文件已生成")

        # 读取验证内容
        wb_out = openpyxl.load_workbook(output_path)
        ws_out = wb_out.active

        check(ws_out.cell(row=1, column=1).value == "菜品名称", "表头保留 '菜品名称'")
        check(ws_out.cell(row=1, column=4).value == "配料", "表头保留 '配料'")

        # 验证填充数据
        check(ws_out.cell(row=2, column=4).value == "T240、B30/80、S4", "第1行配料已填充")
        check(ws_out.cell(row=3, column=4).value == "T265、B30/105、S5", "第2行配料已填充")

        # 验证置信度列
        check(ws_out.cell(row=1, column=5).value == "匹配置信度", "置信度列表头已添加")
        check(ws_out.cell(row=2, column=5).value == "HIGH", "第1行 HIGH")
        check(ws_out.cell(row=3, column=5).value == "LOW_CONFIDENCE", "第2行 LOW_CONFIDENCE")

        # 验证样式保留（表头加粗 + 蓝色背景）
        cell_a1 = ws_out.cell(row=1, column=1)
        check(cell_a1.font.bold is True, "表头加粗保留")
        check(cell_a1.fill.fgColor.rgb in ("004472C4", "FF4472C4"), "表头蓝色背景保留")

        # 验证列宽保留
        check(ws_out.column_dimensions["A"].width == 20, "列宽保留")
        print()

        # ── 2. 校验报告 ──
        print("2. 校验报告")
        low_rows = [
            {"row_index": 2, "product_name": "五黄高纤慢养瓶", "reason": "商品名匹配低于阈值"},
        ]
        rp = write_report(report_path, low_rows)
        check(os.path.exists(rp), "报告文件已生成")

        report_text = Path(rp).read_text(encoding="utf-8")
        check("低置信度行数: 1" in report_text, "报告含低置信度计数")
        check("商品名匹配低于阈值" in report_text, "报告含失败原因")
        print()

        # ── 3. 空报告（全部 HIGH） ──
        print("3. 空报告（全部 HIGH）")
        empty_report = os.path.join(tmpdir, "empty_report.txt")
        write_report(empty_report, [])
        empty_text = Path(empty_report).read_text(encoding="utf-8")
        check("全部行匹配置信度为 HIGH" in empty_text, "空报告提示无需关注")
        print()

        # ── 4. 目标列已存在时的覆盖 ──
        print("4. 目标列覆盖写入")
        out2 = write_result(template_path, output_path, result_df)
        wb2 = openpyxl.load_workbook(out2)
        ws2 = wb2.active
        check(ws2.cell(row=2, column=4).value == "T240、B30/80、S4", "覆盖后数据正确")
        print()

        # ── 5. write_expanded_template 写入空白模板 ──
        print("5. write_expanded_template 写入选项规格模板")
        opt_out = write_expanded_template(opt_template_path, opt_output_path, expanded_df)
        check(os.path.exists(opt_out), "选项输出文件已生成")

        wb_opt_out = openpyxl.load_workbook(opt_output_path)
        ws_opt_out = wb_opt_out.active

        # 验证表头
        check(ws_opt_out.cell(row=1, column=1).value == "商品编码", "表头「商品编码」")
        check(ws_opt_out.cell(row=1, column=3).value == "口味做法组名", "表头「口味做法组名」")

        # 验证数据行
        check(ws_opt_out.cell(row=2, column=1).value == "A001", "第1行 商品编码")
        check(ws_opt_out.cell(row=2, column=3).value == "糖度", "第1行 口味做法组名=糖度")
        check(ws_opt_out.cell(row=2, column=4).value == "七分糖", "第1行 选项名称=七分糖")
        check(ws_opt_out.cell(row=2, column=5).value == 1, "第1行 最少必选=1（整数）")
        check(ws_opt_out.cell(row=2, column=7).value == "是", "第1行 推荐项=是")
        check(ws_opt_out.cell(row=5, column=4).value == "正常冰", "第5行 选项名称=正常冰（温度维度）")

        # 验证行数
        data_row_count = sum(1 for row in range(2, ws_opt_out.max_row + 2)
                            if ws_opt_out.cell(row=row, column=1).value is not None)
        check(data_row_count == 6, f"共 6 行数据（实际 {data_row_count}）")
        print()

        # ── 6. write_expanded_template 保留模板格式 ──
        print("6. write_expanded_template 保留模板格式")
        check(ws_opt_out.column_dimensions["A"].width == 15, "列宽保留")
        cell_h = ws_opt_out.cell(row=1, column=1)
        check(cell_h.font.bold is True, "表头加粗保留")
        # 不相关列未被修改
        check(ws_opt_out.cell(row=1, column=9).value == "其他列", "其他列未受影响")
        print()

        # ── 7. 模板文件不存在 → FileNotFoundError ──
        print("7. write_expanded_template 模板不存在 → FileNotFoundError")
        try:
            write_expanded_template("不存在的模板.xlsx", opt_output_path, expanded_df)
            check(False, "应抛出 FileNotFoundError")
        except FileNotFoundError as e:
            check("模板文件不存在" in str(e), f"报错含「模板文件不存在」")
        print()

    finally:
        for f in [template_path, output_path, report_path,
                  os.path.join(tmpdir, "empty_report.txt"),
                  opt_template_path, opt_output_path]:
            if os.path.exists(f):
                os.remove(f)
        os.rmdir(tmpdir)

    print(f"=== 结果: {passed} passed, {failed} failed ===")
